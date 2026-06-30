import io
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(__file__))
REF  = os.path.join(BASE, "reference-files")

CAT_MAP_PATH    = os.path.join(REF, "daraz_and_cartup_category_mapping.xlsx")
CARTUP_CAT_PATH = os.path.join(REF, "category_mapping.xlsx")
VAR_MATCH_PATH  = os.path.join(REF, "Category_wise_varints.xlsx")

MANUAL_CAT_MAP = {
    '10000860': '4681', '10000866': '4681', '10000887': '4681',
    '10002020': '4660', '10121':    '4690',
}
MYSTERY_CATIDS = {'20000359'}

VARIANT_COLS = [
    'Clothing Materials','Shoe Material','Bag Material','Dial Materials',
    'Strap Materials','Main Materials','Recommended Age','Watch TYespe',
    'Clothing Size','Age Group','Shoe Size','Size','Color','Bedding Size','Model'
]

SECTIONS = [
    ('Basic Information',   'D9E1F2', 15),
    ('Product Attribute',   'E2EFDA', 8),
    ('Product Description', 'FCE4D6', 5),
    ('Service',             'FFF2CC', 4),
    ('Delivery',            'DDEBF7', 4),
    ('Variant Attribute',   'F2F2F2', 11),
    ('Extra',               'EDEDED', 4),
]

# ── openpyxl helpers ───────────────────────────────────────────────────────
def _wb_to_rows(path_or_bytes, sheet_name=None, skip_rows=0):
    """Read xlsx → list of dicts using openpyxl (no pandas)."""
    if isinstance(path_or_bytes, (str, os.PathLike)):
        wb = load_workbook(path_or_bytes, read_only=True, data_only=True)
    else:
        wb = load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows: return []
    header = [str(c).strip() if c is not None else '' for c in rows[skip_rows]]
    result = []
    for row in rows[skip_rows + 1:]:
        d = {header[i]: (_v(row[i]) if i < len(row) else '') for i in range(len(header))}
        result.append(d)
    return result, header

def _wb_sheets(path_or_bytes, skip_rows=0):
    """Read all sheets → dict of sheet_name: list of dicts."""
    if isinstance(path_or_bytes, (str, os.PathLike)):
        wb = load_workbook(path_or_bytes, read_only=True, data_only=True)
    else:
        wb = load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        header = [str(c).strip() if c is not None else '' for c in rows[skip_rows]]
        sheet_rows = []
        for row in rows[skip_rows + 1:]:
            d = {header[i]: (_v(row[i]) if i < len(row) else '') for i in range(len(header))}
            sheet_rows.append(d)
        result[name] = (sheet_rows, header)
    wb.close()
    return result

def _v(x):
    if x is None: return ''
    s = str(x).strip()
    return '' if s == 'nan' else s

# ── Load reference files ───────────────────────────────────────────────────
def _load_references():
    # Category map: Daraz → Cartup
    cat_rows, _ = _wb_to_rows(CAT_MAP_PATH)
    daraz_to_cartup = {}
    for r in cat_rows:
        d = _v(r.get('Daraz Category ID'))
        c = _v(r.get('Cartup Category ID'))
        if d and c:
            daraz_to_cartup[d] = c
    daraz_to_cartup.update(MANUAL_CAT_MAP)

    # Cartup categories: ID → Path + Tags
    cartup_rows, cartup_header = _wb_to_rows(CARTUP_CAT_PATH)
    cartup_to_path = {}
    cartup_to_tags = {}
    tag_col = cartup_header[2] if len(cartup_header) > 2 else ''
    for r in cartup_rows:
        cid  = _v(r.get('Cartup Category ID'))
        path = _v(r.get('Cartup Category Path'))
        tags = _v(r.get(tag_col, ''))
        if cid:
            cartup_to_path[cid] = path
            cartup_to_tags[cid] = tags

    # Variant matching: Cartup catId → applicable variant cols
    var_sheets = _wb_sheets(VAR_MATCH_PATH)
    var_rows, var_header = var_sheets.get('Variations matching', ([], []))
    cat_variant_map = {}
    for col in VARIANT_COLS:
        if col not in var_header: continue
        col_idx = var_header.index(col)
        status_col = var_header[col_idx + 1] if col_idx + 1 < len(var_header) else ''
        for r in var_rows:
            cat_id = _v(r.get(col))
            status = _v(r.get(status_col))
            if cat_id and status.lower() == 'yes':
                cat_variant_map.setdefault(cat_id, set()).add(col)

    return daraz_to_cartup, cartup_to_path, cartup_to_tags, cat_variant_map

def _build_brand_dict(attr_bytes):
    if not attr_bytes: return {}
    brand_dict = {}
    sheets = _wb_sheets(attr_bytes, skip_rows=1)
    for sheet_name, (rows, header) in sheets.items():
        if sheet_name == 'INDEX': continue
        pid_col   = next((c for c in header if 'Product ID' in c), None)
        brand_col = next((c for c in header if 'Brand' in c), None)
        if not pid_col or not brand_col: continue
        for r in rows:
            pid   = _v(r.get(pid_col))
            brand = _v(r.get(brand_col))
            if pid and pid not in brand_dict:
                brand_dict[pid] = brand if brand else 'No Brand'
    return brand_dict

def _clean_highlights(html):
    if not html: return ''
    soup = BeautifulSoup(str(html), 'html.parser')
    for img in soup.find_all('img'): img.decompose()
    items = [f'<li>{li.get_text(strip=True)}</li>' for li in soup.find_all('li') if li.get_text(strip=True)]
    if items: return '<ul>' + ''.join(items) + '</ul>'
    lines = [l for l in soup.get_text(separator='\n', strip=True).split('\n') if l]
    return '<ul>' + ''.join(f'<li>{l}</li>' for l in lines) + '</ul>' if lines else ''

def _clean_description(html):
    if not html: return ''
    soup = BeautifulSoup(str(html), 'html.parser')
    for img in soup.find_all('img'): img.decompose()
    paras = [f'<p>{p.get_text(strip=True)}</p>' for p in soup.find_all(['p','pre']) if p.get_text(strip=True)]
    if paras: return ''.join(paras)
    text = soup.get_text(separator=' ', strip=True)
    return f'<p>{text}</p>' if text else ''

def _parse_variations(combo, applicable):
    result = {c: '' for c in VARIANT_COLS}
    if not combo: return result
    parts = [p.strip() for p in combo.split(',', 1)]
    color    = parts[0]
    size_val = parts[1] if len(parts) > 1 else 'Yes'
    if 'Color' in applicable: result['Color'] = color
    if size_val:
        if 'Shoe Size' in applicable:       result['Shoe Size'] = size_val
        elif 'Clothing Size' in applicable: result['Clothing Size'] = size_val
        elif 'Bedding Size' in applicable:  result['Bedding Size'] = size_val
        elif 'Size' in applicable:          result['Size'] = size_val
        else:                               result['Size'] = size_val
    return result

def _build_excel(output_rows):
    cols = list(output_rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = 'product'

    sec_row = []
    for name_s, color, span in SECTIONS:
        sec_row.extend([name_s] + [''] * (span - 1))
    ws.append(sec_row)

    col_pos = 1
    for name_s, color, span in SECTIONS:
        for i in range(span):
            c = ws.cell(row=1, column=col_pos + i)
            c.fill = PatternFill('solid', start_color=color)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_pos, end_row=1, end_column=col_pos + span - 1)
        col_pos += span

    ws.append(cols)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci)
        c.fill = PatternFill('solid', start_color='D9D9D9')
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for r in output_rows:
        ws.append([r.get(c, '') for c in cols])

    report_col_idx = cols.index('Report') + 1
    for row_i in range(3, len(output_rows) + 3):
        cell = ws.cell(row=row_i, column=report_col_idx)
        v = str(cell.value or '')
        if v == 'OK':
            cell.fill = PatternFill('solid', start_color='E2EFDA')
            cell.font = Font(color='375623')
        elif 'Mystery' in v:
            cell.fill = PatternFill('solid', start_color='FCE4D6')
            cell.font = Font(color='9C0006', bold=True)
        elif 'Manually' in v:
            cell.fill = PatternFill('solid', start_color='FFF2CC')
            cell.font = Font(color='7F6000')
        elif 'No category' in v:
            cell.fill = PatternFill('solid', start_color='FFDCE1')
            cell.font = Font(color='9C0006')

    for ci in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions[get_column_letter(report_col_idx)].width = 45
    ws.freeze_panes = 'A3'
    ws.row_dimensions[2].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def process_daraz_files(price_b, basic_b, weight_b, skuimg_b, attr_b=None):
    daraz_to_cartup, cartup_to_path, cartup_to_tags, cat_variant_map = _load_references()
    brand_dict = _build_brand_dict(attr_b)

    price_rows,   _ = _wb_to_rows(price_b,  sheet_name='template')
    basic_rows,   bh = _wb_to_rows(basic_b,  sheet_name='template')
    freight_rows, _ = _wb_to_rows(weight_b, sheet_name='template')
    skuimg_rows,  _ = _wb_to_rows(skuimg_b, sheet_name='template')

    # Build lookup dicts
    basic_dict   = {}
    for r in basic_rows:
        pid = _v(r.get('Product ID'))
        if pid and pid not in basic_dict:
            basic_dict[pid] = r

    freight_dict = {}
    for r in freight_rows:
        pid = _v(r.get('Product ID'))
        if pid and pid not in freight_dict:
            freight_dict[pid] = r

    skuimg_dict = {}
    for r in skuimg_rows:
        sku = _v(r.get('SellerSKU'))
        if sku and sku not in skuimg_dict:
            skuimg_dict[sku] = r

    warranty_type_col = '*Warranty Type' if '*Warranty Type' in bh else 'Warranty Type'

    output_rows = []
    for row in price_rows:
        pid   = _v(row.get('Product ID'))
        cat   = _v(row.get('catId'))
        name  = _v(row.get('*Product Name(English)'))
        sku   = _v(row.get('SellerSKU'))
        combo = _v(row.get('Variations Combo'))

        brand = brand_dict.get(pid, 'No Brand')
        if not brand: brand = 'No Brand'

        is_mystery  = cat in MYSTERY_CATIDS
        report_note = ''

        if is_mystery:
            cartup_id = cartup_path = tags = ''
            report_note = 'Mystery/Surprise Box — no category. Manual review needed.'
        else:
            cartup_id   = daraz_to_cartup.get(cat, '')
            cartup_path = cartup_to_path.get(cartup_id, '')
            tags        = cartup_to_tags.get(cartup_id, '')
            if not cartup_id:
                report_note = f'No category match for Daraz catId={cat}'
            elif cat in MANUAL_CAT_MAP:
                report_note = f'Manually mapped (Daraz catId={cat})'
            else:
                report_note = 'OK'

        applicable = cat_variant_map.get(cartup_id, set())

        b = basic_dict.get(pid, {})
        highlights  = _clean_highlights(_v(b.get('*Highlights')))
        description = _clean_description(_v(b.get('Main Description')))

        if not highlights and not description:
            highlights  = f'<ul><li>{name}</li></ul>'
            description = f'<p>{name}</p>'
        elif not highlights:
            text = BeautifulSoup(description, 'html.parser').get_text(strip=True)
            highlights = f'<ul><li>{name}</li><li>{text[:200]}</li></ul>'
        elif not description:
            text = BeautifulSoup(highlights, 'html.parser').get_text(separator=' ', strip=True)
            description = f'<p>{name}. {text[:300]}</p>'

        f = freight_dict.get(pid, {})
        s = skuimg_dict.get(sku, {})

        def img(i):
            k = '*Product Images1' if i == 1 else f'Product Images{i}'
            return _v(b.get(k))

        vr = _parse_variations(combo, applicable)
        warranty_policy = _v(b.get('Warranty Policy'))

        output_rows.append({
            '**Category Id':            cartup_id,
            '**Name (English)':         name,
            'Name (Bengali)':           name,
            '**Product Image 1':        img(1), 'Product Image 2': img(2),
            'Product Image 3':          img(3), 'Product Image 4': img(4),
            'Product Image 5':          img(5), 'Product Image 6': img(6),
            'Product Image 7':          img(7), 'Product Image 8': img(8),
            'VideoUrl':                 '',
            '**Brand':                  brand,
            '**Unit':                   'pcs',
            'Tags':                     tags,
            'Clothing Materials':       vr['Clothing Materials'],
            'Shoe Material':            vr['Shoe Material'],
            'Bag Material':             vr['Bag Material'],
            'Dial Materials':           vr['Dial Materials'],
            'Strap Materials':          vr['Strap Materials'],
            'Recommended Age':          vr['Recommended Age'],
            'Watch Type':               vr['Watch TYespe'],
            'Main Materials':           vr['Main Materials'],
            'Highlights(English)':      highlights,
            'Highlights(Bengali)':      highlights,
            'Description (Bengali)':    description,
            'Description (English)':    description,
            "What's in the box":        f'1* {name}',
            'Warranty Policy(English)': warranty_policy,
            'Warranty Policy(Bangla)':  warranty_policy,
            'Warranty Type':            _v(b.get(warranty_type_col)),
            'Warranty Period':          _v(b.get('Warranty')),
            '**Package Weight (kg)':    _v(f.get('*Package Weight (kg)')),
            '**Package Length(cm)':     _v(f.get('*Package Length (cm)')),
            '*Package Width (cm)':      _v(f.get('*Package Width (cm)')),
            '*Package Height(cm)':      _v(f.get('*Package Height (cm)')),
            'Clothing Size':            vr['Clothing Size'],
            'Color':                    vr['Color'],
            'Model':                    vr['Model'],
            'Age Group':                vr['Age Group'],
            'Size':                     vr['Size'],
            'Shoe Size':                vr['Shoe Size'],
            'Bedding Size':             vr['Bedding Size'],
            '**Seller SKU':             sku,
            '**Parent Sku':             pid,
            '*Variant Image':           _v(s.get('Images1')),
            '**Current Stock Qty':      _v(row.get('*Quantity')),
            'status':                   _v(row.get('status')),
            'Cartup Category Path':     cartup_path,
            'Variations Combo':         combo,
            'Report':                   report_note,
        })

    return _build_excel(output_rows)

def process_manual_upload(input_bytes):
    """
    Reads a manual product Excel file, cleans cell values, and returns
    the same data as a clean Excel. Full AI processing is done client-side
    via manualProcessor.js — this endpoint is kept as a server-side fallback.
    """
    wb = load_workbook(io.BytesIO(input_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'product'

    for ri, row in enumerate(rows):
        out_ws.append([_v(c) for c in row])
        if ri == 0:
            for ci in range(1, len(row) + 1):
                cell = out_ws.cell(row=1, column=ci)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', start_color='D9D9D9')
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for ci in range(1, (len(rows[0]) if rows else 1) + 1):
        out_ws.column_dimensions[get_column_letter(ci)].width = 22
    out_ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue()
